"""
Universal File Reader — extracts content from any file type.

Supports: text files, PDF, DOCX, CSV, Excel, images (base64), zip/folders, JSON.
Returns a uniform dict that agents can consume directly.

Used by:
  - POST /chat/stream/with-files (inject file content as context)
  - Shell agent (when asked to read an uploaded file)
  - Vision agent (when a file is uploaded for image analysis)
"""

import asyncio
import base64
import csv
import io
import json
import zipfile
from pathlib import Path
from typing import Any, Dict, Optional

import structlog
logger = structlog.get_logger(__name__)

TEXT_EXT = {
    ".txt", ".md", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css", ".sh",
    ".bash", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".env", ".sql", ".r",
    ".rb", ".go", ".rs", ".java", ".cpp", ".c", ".h", ".cs", ".php", ".swift",
    ".kt", ".scala", ".lua", ".pl", ".xml",
}
IMG_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".tiff"}


async def _read_text(p: Path, max_chars: int = 50_000) -> dict:
    try:
        text = p.read_text(encoding="utf-8", errors="replace")
        return {
            "type": "text",
            "content": text[:max_chars],
            "truncated": len(text) > max_chars,
            "chars": len(text),
        }
    except Exception as e:
        return {"type": "text", "error": str(e)}


async def _read_pdf(p: Path) -> dict:
    try:
        from pdfminer.high_level import extract_text
        text = await asyncio.to_thread(extract_text, str(p))
        return {"type": "pdf", "content": text[:50_000], "truncated": len(text) > 50_000}
    except ImportError:
        try:
            proc = await asyncio.create_subprocess_shell(
                f"pdftotext \"{p}\" -",
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            out, _ = await proc.communicate()
            text = out.decode("utf-8", "replace")
            return {"type": "pdf", "content": text[:50_000], "truncated": len(text) > 50_000}
        except Exception as e:
            return {"type": "pdf", "error": str(e)}
    except Exception as e:
        return {"type": "pdf", "error": str(e)}


async def _read_docx(p: Path) -> dict:
    try:
        from docx import Document
        doc = await asyncio.to_thread(Document, str(p))
        text = "\n".join(para.text for para in doc.paragraphs)
        return {"type": "docx", "content": text[:50_000], "truncated": len(text) > 50_000}
    except ImportError:
        return {"type": "docx", "error": "pip install python-docx"}
    except Exception as e:
        return {"type": "docx", "error": str(e)}


async def _read_csv(p: Path) -> dict:
    try:
        rows = []
        with p.open(encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for i, row in enumerate(reader):
                if i >= 1000:
                    break
                rows.append(dict(row))
        return {
            "type": "csv",
            "rows": rows,
            "row_count": len(rows),
            "columns": list(rows[0].keys()) if rows else [],
        }
    except Exception as e:
        return {"type": "csv", "error": str(e)}


async def _read_excel(p: Path) -> dict:
    try:
        import openpyxl
        wb = await asyncio.to_thread(openpyxl.load_workbook, str(p), read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames[:5]:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 500:
                    break
                rows.append([str(c) if c is not None else "" for c in row])
            sheets[name] = rows
        return {"type": "excel", "sheets": sheets}
    except ImportError:
        return {"type": "excel", "error": "pip install openpyxl"}
    except Exception as e:
        return {"type": "excel", "error": str(e)}


async def _read_image(p: Path) -> dict:
    data = p.read_bytes()
    suffix = p.suffix.lower().lstrip(".")
    mime_map = {
        "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
        "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp",
    }
    return {
        "type": "image",
        "base64": base64.b64encode(data).decode(),
        "mime_type": mime_map.get(suffix, "image/png"),
        "size": len(data),
    }


async def _read_zip(p: Path, temp: Path) -> dict:
    extract_to = temp / p.stem
    extract_to.mkdir(exist_ok=True)
    with zipfile.ZipFile(p, "r") as zf:
        zf.extractall(extract_to)
    files = []
    for i, fp in enumerate(extract_to.rglob("*")):
        if i >= 100:
            break
        if fp.is_file():
            content = await extract_any_file(fp, temp)
            files.append({"path": str(fp.relative_to(extract_to)), "content": content})
    return {"type": "zip", "file_count": len(files), "files": files}


async def extract_any_file(path: Path, temp_dir: Optional[Path] = None) -> Dict[str, Any]:
    """Route any file to the appropriate extractor. Returns a uniform content dict."""
    ext = path.suffix.lower()

    # ── Format-specific handlers take priority over generic text fallback ────
    if ext == ".pdf":
        return await _read_pdf(path)
    if ext in {".docx", ".doc"}:
        return await _read_docx(path)
    if ext == ".csv":
        return await _read_csv(path)
    if ext in {".xlsx", ".xls"}:
        return await _read_excel(path)
    if ext in IMG_EXT:
        return await _read_image(path)
    if ext == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            return {"type": "json", "content": json.dumps(data, indent=2)[:20_000]}
        except Exception as e:
            return {"type": "json", "error": str(e)}
    if ext == ".zip" and temp_dir:
        return await _read_zip(path, temp_dir)

    # ── Text files (code, markdown, config, etc.) ───────────────────────────
    if ext in TEXT_EXT or (path.stat().st_size < 100_000 and ext not in IMG_EXT):
        return await _read_text(path)

    # Final fallback: try as text
    return await _read_text(path)

